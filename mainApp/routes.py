from mainApp import app
from mainApp import db
from mainApp.forms import RegisterForm, LoginForm, StatusForm, ProductForm, EventForm, ProductOpenStatusForm, ProductCloseStatusForm, ProductAutoStatusForm, ProductEditForm, EventStartForm, EventCloseForm, SetdateRange
from mainApp.models.user import User
from mainApp.models.product import Product
from mainApp.models.event import Event
from mainApp.models.status import Status
from flask_login import login_required, logout_user, login_user, current_user
from flask import render_template, request, redirect, url_for, flash, send_file
from datetime import datetime, timedelta
from openpyxl import Workbook
import string
from sqlalchemy import or_, and_, func
from functools import wraps
from flask_babel import gettext
import re
import matplotlib.pyplot as plt


def admin_check(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if current_user.is_anonymous:
            return redirect(url_for('login_page'))
        elif current_user.role == "admin":
            print("you are admin")
        else:
            flash(f'You are not admin!', category='danger')
            return render_template('404.html')
        return func(*args, **kwargs)
    return wrapper


def openStatusesCounter():
    openEventCounter = 0
    activEventsList = Event.query.filter(Event.endDate == None)
    for activEventList in activEventsList:
        openEventCounter += 1
    return openEventCounter


@app.route('/')
@app.route('/home')
def home_page():
    openStatuses = openStatusesCounter()
    prodStart = gettext('start produkcji')

    return render_template('home.html', prodStart=prodStart, openStatuses=openStatuses)


@app.route('/register', methods=['GET', 'POST'])
def register_page():
    form = RegisterForm()
    if form.validate_on_submit():
        user_to_create = User(
            username=form.username.data, email_address=form.email_address.data, role=form.role.data)
        user_to_create.set_password(form.password1.data)
        db.session.add(user_to_create)
        db.session.commit()
        flash(
            f'Success! usser created: {user_to_create.username}', category='success')
        login_user(user_to_create)
        return redirect(url_for('home_page'))
    if form.errors != {}:  # validation errors
        for err_msg in form.errors.values():
            flash(
                f'There was an error with creating a user: {err_msg}', category='danger')

    return render_template('registerForm.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    form = LoginForm()
    if form.validate_on_submit():
        attempted_user = User.query.filter(
            User.username == form.username.data).first()
        if attempted_user and attempted_user.check_password(form.password.data):
            login_user(attempted_user)
            flash(
                f'Success! You are logged in as: {attempted_user.username}', category='success')
            return redirect(url_for('home_page'))
        else:
            flash(f'User name or password incorrect!', category='danger')

    return render_template('loginForm.html', form=form)


@app.route('/logout')
@login_required
def logout_page():
    logout_user()
    flash(f'You have been logged out!', category='info')
    return redirect(url_for('home_page'))


@app.route('/status', methods=['GET', 'POST'])
@login_required
@admin_check
def status_page():
    openStatuses = openStatusesCounter()
    form = StatusForm()
    if form.validate_on_submit():
        status_to_create = Status(
            form.statusCode.data, form.statusName.data, form.production.data)
        db.session.add(status_to_create)
        db.session.commit()
        flash(
            f'Success! status added: {status_to_create.statusName}', category='success')
        return redirect(url_for('home_page'))
    if form.errors != {}:  # validation errors
        for err_msg in form.errors.values():
            flash(f'Status incorrect!: {err_msg}', category='danger')

    return render_template('statusForm.html', form=form, openStatuses=openStatuses)


@app.route('/product', methods=['GET', 'POST'])
@login_required
@admin_check
def product_page():

    form = ProductForm()

    if form.validate_on_submit():
        product_to_create = Product(form.modelCode.data, form.modelName.data,
                                    form.orderStatus.data, int(form.startDate.data.timestamp()), int(form.executionDate.data.timestamp()))
        db.session.add(product_to_create)
        db.session.commit()
        flash(
            f'Success! product added: {product_to_create.modelCode}', category='success')
    if form.errors != {}:  # validation errors
        for err_msg in form.errors.values():
            flash(f'Product incorrect!: {err_msg}', category='danger')
    openStatuses = openStatusesCounter()
    return render_template('productForm.html', form=form, openStatuses=openStatuses)


@app.route('/event', methods=['GET', 'POST'])
@login_required
@admin_check
def event_page():

    EventForm.activProductList.clear()
    EventForm.idStatusList.clear()
    activProductListDB = Product.query.filter(Product.orderStatus == 'Open')
    for row in activProductListDB:
        EventForm.activProductList.append(
            (row.id, str(row.modelCode) + " - " + row.modelName))
    idStatusListDB = Status.query.all()
    for row in idStatusListDB:
        EventForm.idStatusList.append(
            (row.id, str(row.statusCode) + " - " + row.statusName))
    form = EventForm()

    if form.validate_on_submit():
        startDate = form.startDate.data
        endDate = form.endDate.data
        event_to_create = Event(int(form.idProd.data), int(form.idStatus.data), int(form.startDate.data.timestamp()), int(
            form.endDate.data.timestamp()), form.okCounter.data, form.nokCounter.data, int(form.userID.data))
        db.session.add(event_to_create)
        db.session.commit()
        flash(
            f'Success! event added: {event_to_create}', category='success')
    if form.errors != {}:  # validation errors
        for err_msg in form.errors.values():
            flash(f'Product incorrect!: {err_msg}', category='danger')
    openStatuses = openStatusesCounter()
    return render_template('eventForm.html', form=form, openStatuses=openStatuses)


@ app.route('/eventStartStop', methods=('GET', 'POST'))
@ login_required
def event_start_stop_page():
    eventStartForm = EventStartForm()
    eventCloseForm = EventCloseForm()
    if request.method == 'POST':
        if request.form['idEvent'] != "None":
            okCounter = request.form['okCounter']
            nokCounter = request.form['nokCounter']
            eventID = request.form['idEvent']
            event_to_close = Event.query.get(eventID)

            now = datetime.today()
            now = now.timestamp()
            now = int(now)
            event_to_close.okCounter = okCounter
            event_to_close.nokCounter = nokCounter
            event_to_close.endDate = now

            db.session.add(event_to_close)
            db.session.commit()
            flash(
                f'Success! Event closed: {eventID}', category='success')
            if request.form['production'] == "Finish":
                idProd = request.form['idProd']
                product_to_close = Product.query.get(idProd)
                product_to_close.orderStatus = "Finished"
                db.session.commit()
                flash(
                    f'Success! Product Finished: {product_to_close.modelCode}', category='success')
        else:
            idProd = request.form['idProd']
            print(idProd)
            idStatus = request.form['idStatus']
            print(idStatus)
            idUser = request.form['idUser']
            print(idUser)
            today = datetime.today()
            today = today.timestamp()
            today = int(today)
            print(today)
            event_to_create = Event(idProd=idProd, idStatus=idStatus, startDate=today, endDate=None, nokCounter=None,
                                    okCounter=None,  userID=idUser)
            db.session.add(event_to_create)
            db.session.commit()
            flash(
                f'Success! Event Start: {idProd} - {idStatus}', category='success')
    now = datetime.today()
    now = now.timestamp()
    now = int(now)
    print(now)
    openProductList = Product.query.filter(or_(Product.orderStatus == "Open", and_(
        Product.orderStatus == "Auto", Product.startDate <= now, Product.executionDate >= now)))
    statusList = Status.query.all()
    openEventList = Event.query.filter(Event.endDate == None)
    openStatuses = openStatusesCounter()

    return render_template('eventStartForm.html', openProductList=openProductList, statusList=statusList, openEventList=openEventList, eventStartForm=eventStartForm, eventCloseForm=eventCloseForm, openStatuses=openStatuses)


@ app.route('/eventAllStop', methods=('GET', 'POST'))
@ login_required
def event_all_stop_page():
    now = datetime.today()
    now = now.timestamp()
    now = int(now)
    activEventList = Event.query.filter(Event.endDate == None)
    for column in activEventList:
        print(column.id)
        column.endDate = now
    db.session.commit()
    return redirect((url_for('event_start_stop_page')))


@app.route('/productTable', methods=['GET', 'POST'])
@login_required
def product_table_page():
    productCloseStatusForm = ProductCloseStatusForm()
    productOpenStatusForm = ProductOpenStatusForm()
    productAutoStatusForm = ProductAutoStatusForm()
    productEditForm = ProductEditForm()
    if request.method == 'POST':
        productID = request.form.get('productID')
        newStatus = request.form.get('newStatus')
        if newStatus == "Edit":
            modelCode = request.form.get('modelCode')
            modelName = request.form.get('modelName')
            orderStatus = request.form.get('orderStatus')
            startDate = request.form.get('startDate')
            executionDate = request.form.get('executionDate')
            startDateDate = datetime.strptime(startDate, "%Y-%m-%dT%H:%M")
            executionDateDate = datetime.strptime(
                executionDate, "%Y-%m-%dT%H:%M")
            startDateDate = startDateDate.timestamp()
            executionDateDate = executionDateDate.timestamp()
            product = Product.query.get(productID)
            oldName = product.modelName
            product.modelCode = modelCode
            product.modelName = modelName
            product.orderStatus = orderStatus
            product.startDate = startDateDate
            product.executionDate = executionDateDate
            db.session.commit()
            flash(
                f"Edited product from {oldName} to {modelName}", category='success')
        else:
            product = Product.query.get(productID)
            oldStatus = product.orderStatus
            product.orderStatus = newStatus
            db.session.commit()
            flash(
                f"Changed product status from {oldStatus} to {newStatus}", category='success')

    # products = Product.query.filter(Product.orderStatus == 'Finished')
    products = Product.query.all()
    for product in products:
        product.delta = "update"
        if product.startDate:
            product.delta = round(
                (product.executionDate - product.startDate)/86400, 1)
            product.startDate = datetime.fromtimestamp(product.startDate)
            product.executionDate = datetime.fromtimestamp(
                product.executionDate)
        else:
            product.delta = "update"
    openStatuses = openStatusesCounter()
    return render_template('productTable.html', products=products, productOpenStatusForm=productOpenStatusForm, productCloseStatusForm=productCloseStatusForm, productAutoStatusForm=productAutoStatusForm, productEditForm=productEditForm, openStatuses=openStatuses)


@app.route('/productFinishedTable', methods=['GET', 'POST'])
@login_required
def product_finished_table_page():
    products = Product.query.filter(Product.orderStatus == 'Finished')
    for product in products:
        if product.startDate:
            product.startDate = datetime.fromtimestamp(product.startDate)
            product.executionDate = datetime.fromtimestamp(
                product.executionDate)

    openStatuses = openStatusesCounter()
    return render_template('productFinishedTable.html', products=products, openStatuses=openStatuses)


@app.route('/eventTable', methods=['GET', 'POST'])
@login_required
@admin_check
def event_table_page():
    startDateTimestamp = datetime.now()
    endDateTimestamp = datetime.now()

    dateRangeMin = 0
    dateRangeMax = 0

    form = SetdateRange()

    startDateTimestamp = datetime.now() - timedelta(days=1)
    startDateTimestamp = startDateTimestamp.strftime('%Y-%m-%dT00:00')

    endDateTimestamp = datetime.now()
    endDateTimestamp = endDateTimestamp.strftime('%Y-%m-%dT23:59')

    form.startDate.default = startDateTimestamp
    form.endDate.default = endDateTimestamp

    if form.validate_on_submit():
        startDateTimestamp = form.startDate.data
        endDateTimestamp = form.endDate.data
        dateRangeMin = startDateTimestamp.timestamp()
        dateRangeMax = endDateTimestamp.timestamp()
        print(dateRangeMin)
        print(dateRangeMax)
    # results = db.session.query(Status, User, Product, Event).filter(
        # Event.startDate, Event.startDate, Event.userID == User.id, Event.idProd == Product.id, Event.idStatus == Status.id)
    results = db.session.query(Status, User, Product, Event).filter(
        Event.startDate >= dateRangeMin, Event.startDate <= dateRangeMax,  Event.userID == User.id, Event.idProd == Product.id, Event.idStatus == Status.id)
    print(dateRangeMin)
    dateRangeMin = startDateTimestamp
    print(dateRangeMax)
    dateRangeMax = endDateTimestamp

    finalEventTable = []
    for status, user, product, event in results:
        finalEvent = {}
        finalEvent["id"] = event.id
        finalEvent["modelCode"] = product.modelCode
        finalEvent["modelName"] = product.modelName
        finalEvent["username"] = user.username
        finalEvent["statusName"] = status.statusName
        finalEvent["okCounter"] = event.okCounter
        finalEvent["nokCounter"] = event.nokCounter
        finalEvent["startDate"] = datetime.fromtimestamp(event.startDate)
        if event.endDate:
            finalEvent["endDate"] = datetime.fromtimestamp(event.endDate)
            finalEvent["delta"] = round(
                (event.endDate - event.startDate)/3600, 1)
        else:
            finalEvent["endDate"] = "in progress"
            finalEvent["delta"] = "in progress"
        finalEventTable.append(finalEvent)

    # for resfinalEvent in finalEventTable:
    #     print(resfinalEvent)

    if form.validate_on_submit():
        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet("test")
        ws.title = "New Title"
        ws.sheet_properties.tabColor = "1072BA"
        ws['A1'] = "id"
        ws['B1'] = "modelCode"
        ws['C1'] = "modelName"
        ws['D1'] = "username"
        ws['E1'] = "statusName"
        ws['F1'] = "okCounter"
        ws['G1'] = "nokCounter"
        ws['H1'] = "startDate"
        ws['I1'] = "endDate"
        ws['J1'] = "delta[h]"

        # alphabet = list(string.ascii_lowercase)
        # print(alphabet)
        row_counter = 1
        for resfinalEvent in finalEventTable:
            row_counter += 1
            ws['A'+str(row_counter)] = resfinalEvent["id"]
            ws['B'+str(row_counter)] = resfinalEvent["modelCode"]
            ws['C'+str(row_counter)] = resfinalEvent["modelName"]
            ws['D'+str(row_counter)] = resfinalEvent["username"]
            ws['E'+str(row_counter)] = resfinalEvent["statusName"]
            ws['F'+str(row_counter)] = resfinalEvent["okCounter"]
            ws['G'+str(row_counter)] = resfinalEvent["nokCounter"]
            ws['H'+str(row_counter)] = resfinalEvent["startDate"]
            ws['I'+str(row_counter)] = resfinalEvent["endDate"]
            ws['J'+str(row_counter)] = resfinalEvent["delta"]
            # print(row_counter)
        # output_file_name = "output/raport - " + str(date_now) + ".xls"
        output_file_name = "output/raport.xls"
        wb.save(output_file_name)
        flash(
            f'Export complete! File name: {output_file_name}', category='success')
    openStatuses = openStatusesCounter()

    return render_template('eventTable.html', finalEventTable=finalEventTable, form=form, endDateTimestamp=endDateTimestamp, startDateTimestamp=startDateTimestamp, openStatuses=openStatuses)


@app.route('/download')
@login_required
@admin_check
def download_report_page():
    path = "../output/raport.xls"
    try:
        return send_file(path, as_attachment=True)
    except:
        flash(f'File download error!', category='danger')
        return render_template('404.html')


@app.route('/activeProduct', methods=['GET', 'POST'])
@login_required
def active_product_page():
    productExist = False
    if request.method == 'POST':
        modelCode = request.form.get('modelCode')
        print(modelCode)
        products = Product.query.all()
        for product in products:
            if product.orderStatus != "Finished":
                product.orderStatus = "Close"
            if product.modelCode == modelCode:
                product.orderStatus = "Open"
                productExist = True
        db.session.commit()

        if not productExist:

            modelCodeRegexp = re.search(
                "[0-9]{4}\/[0-9]{2}\/[0-9]{2}\/ZP\/[0-9]+", modelCode)
            if not modelCodeRegexp:
                flash(
                    f'nameCode validation error: {modelCode}', category='danger')
                return redirect((url_for('event_start_stop_page')))

            product_to_create = Product(modelCode, "added by operator",
                                        "Open", None, None)
            db.session.add(product_to_create)
            db.session.commit()
            flash(
                f'Success! product added by operator: {product_to_create.modelCode}', category='success')

    return redirect((url_for('event_start_stop_page')))


@app.route('/productSummary', methods=['GET', 'POST'])
@login_required
def product_summary_page():
    if request.method == 'POST':
        idProd = request.form.get('productID')

        # modelCode = request.form.get('modelCode')
        results = db.session.query(Status, Event).with_entities(Event.id, Event.idStatus,  func.sum(Event.startDate).label("startDate"),  func.sum(Event.endDate).label("endDate"), func.sum(
            Event.endDate - Event.startDate).label("delta"), func.sum(Event.okCounter).label("okCounter"),  func.sum(Event.nokCounter).label("nokCounter"), Status.statusName, Status.production).filter(Event.idProd == idProd, Event.idStatus == Status.id).group_by(Event.idStatus).all()
        statuses = Status.query.all()
        # evetns = Event.query.filter(Event.idProd == idProd)

        finalResultsTable = []
        for result in results:
            finalEvent = {}

            finalEvent["idStatus"] = result.idStatus
            finalEvent["statusName"] = result.statusName
            finalEvent["production"] = result.production
            finalEvent["okCounter"] = result.okCounter
            finalEvent["nokCounter"] = result.nokCounter
            if result.endDate:
                finalEvent["delta"] = round(
                    (result.endDate - result.startDate)/3600, 3)
            else:
                finalEvent["delta"] = "in progress"
            finalResultsTable.append(finalEvent)

        plt.figure(figsize=(8, 5))
        values = []
        labels = []

        for finalResultTable in finalResultsTable:

            labels.append(finalResultTable["statusName"])
            values.append(finalResultTable["delta"])


        plt.pie(values, labels=labels, autopct='%.2f %%')

        plt.title('Summary Chart')

        plt.savefig('mainApp/static/new_plot.png')

        openStatuses = openStatusesCounter()

    return render_template('productSummaryTable.html', finalResultsTable=finalResultsTable, openStatuses=openStatuses)


@ app.route('/help')
def help():
    return'help', 400


@ app.errorhandler(404)
def not_found(e):
    flash(f'404!', category='danger')
    return render_template('404.html')
