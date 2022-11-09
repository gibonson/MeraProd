from mainApp import app
from mainApp import db
from mainApp.models.user import User
from mainApp.models.product import Product
from mainApp.models.status import Status
from mainApp.models.event import Event
from mainApp.auth.auth import admin_check, login_required
from mainApp.events.forms import EventForm, EventStartForm, EventCloseForm, SetdateRange
from mainApp.routes import render_template, flash, request
from sqlalchemy import or_, and_
from datetime import datetime, timedelta
from openpyxl import Workbook



def openEventsCounter():
    openEventCounter = 0
    activEventsList = Event.query.filter(Event.endDate == None)
    for activEventList in activEventsList:
        openEventCounter += 1
    return openEventCounter


@app.route('/event', methods=['GET', 'POST'])
@login_required
@admin_check
def event_page():

    EventForm.activProductList.clear()
    EventForm.idStatusList.clear()
    activProductListDB = Product.query.filter(Product.orderStatus == 'Open')
    for row in activProductListDB:
        EventForm.activProductList.append(
            (row.id, str(row.modelCode) + " - " + row.modelName))
    idStatusListDB = Status.query.all()
    for row in idStatusListDB:
        EventForm.idStatusList.append(
            (row.id, str(row.statusCode) + " - " + row.statusName))
    form = EventForm()

    if form.validate_on_submit():
        startDate = form.startDate.data
        endDate = form.endDate.data
        event_to_create = Event(int(form.idProd.data), int(form.idStatus.data), int(form.startDate.data.timestamp()), int(
            form.endDate.data.timestamp()), form.okCounter.data, form.nokCounter.data, int(form.userID.data))
        db.session.add(event_to_create)
        db.session.commit()
        flash(
            f'Success! event added: {event_to_create}', category='success')
    if form.errors != {}:  # validation errors
        for err_msg in form.errors.values():
            flash(f'Product incorrect!: {err_msg}', category='danger')
    openEvents = openEventsCounter()
    return render_template('eventForm.html', form=form, openEvents=openEvents)


@ app.route('/eventStartStop', methods=('GET', 'POST'))
@ login_required
def event_start_stop_page():
    eventStartForm = EventStartForm()
    eventCloseForm = EventCloseForm()
    if request.method == 'POST':
        if request.form['idEvent'] != "None":
            okCounter = request.form['okCounter']
            nokCounter = request.form['nokCounter']
            eventID = request.form['idEvent']
            event_to_close = Event.query.get(eventID)

            now = datetime.today()
            now = now.timestamp()
            now = int(now)
            event_to_close.okCounter = okCounter
            event_to_close.nokCounter = nokCounter
            event_to_close.endDate = now

            db.session.add(event_to_close)
            db.session.commit()
            flash(
                f'Success! Event closed: {eventID}', category='success')
            if request.form['production'] == "Finish":
                idProd = request.form['idProd']
                product_to_close = Product.query.get(idProd)
                product_to_close.orderStatus = "Finished"
                db.session.commit()
                flash(
                    f'Success! Product Finished: {product_to_close.modelCode}', category='success')
        else:
            idProd = request.form['idProd']
            print(idProd)
            idStatus = request.form['idStatus']
            print(idStatus)
            idUser = request.form['idUser']
            print(idUser)
            today = datetime.today()
            today = today.timestamp()
            today = int(today)
            print(today)
            event_to_create = Event(idProd=idProd, idStatus=idStatus, startDate=today, endDate=None, nokCounter=None,
                                    okCounter=None,  userID=idUser)
            db.session.add(event_to_create)
            db.session.commit()
            flash(
                f'Success! Event Start: {idProd} - {idStatus}', category='success')
    now = datetime.today()
    now = now.timestamp()
    now = int(now)
    print(now)
    openProductList = Product.query.filter(or_(Product.orderStatus == "Open", and_(
        Product.orderStatus == "Auto", Product.startDate <= now, Product.executionDate >= now)))
    statusList = Status.query.all()
    openEventList = Event.query.filter(Event.endDate == None)
    openEvents = openEventsCounter()

    return render_template('eventStartForm.html', openProductList=openProductList, statusList=statusList, openEventList=openEventList, eventStartForm=eventStartForm, eventCloseForm=eventCloseForm, openEvents=openEvents)


@ app.route('/eventAllStop', methods=('GET', 'POST'))
@ login_required
def event_all_stop_page():
    now = datetime.today()
    now = now.timestamp()
    now = int(now)
    activEventList = Event.query.filter(Event.endDate == None)
    for column in activEventList:
        print(column.id)
        column.endDate = now
    db.session.commit()
    return redirect((url_for('event_start_stop_page')))


@app.route('/eventTable', methods=['GET', 'POST'])
@login_required
@admin_check
def event_table_page():
    startDateTimestamp = datetime.now()
    endDateTimestamp = datetime.now()

    dateRangeMin = 0
    dateRangeMax = 0

    form = SetdateRange()

    startDateTimestamp = datetime.now() - timedelta(days=1)
    startDateTimestamp = startDateTimestamp.strftime('%Y-%m-%dT00:00')

    endDateTimestamp = datetime.now()
    endDateTimestamp = endDateTimestamp.strftime('%Y-%m-%dT23:59')

    form.startDate.default = startDateTimestamp
    form.endDate.default = endDateTimestamp

    if form.validate_on_submit():
        startDateTimestamp = form.startDate.data
        endDateTimestamp = form.endDate.data
        dateRangeMin = startDateTimestamp.timestamp()
        dateRangeMax = endDateTimestamp.timestamp()
        print(dateRangeMin)
        print(dateRangeMax)
    # results = db.session.query(Status, User, Product, Event).filter(
        # Event.startDate, Event.startDate, Event.userID == User.id, Event.idProd == Product.id, Event.idStatus == Status.id)
    results = db.session.query(Status, User, Product, Event).filter(
        Event.startDate >= dateRangeMin, Event.startDate <= dateRangeMax,  Event.userID == User.id, Event.idProd == Product.id, Event.idStatus == Status.id)
    print(dateRangeMin)
    dateRangeMin = startDateTimestamp
    print(dateRangeMax)
    dateRangeMax = endDateTimestamp

    finalEventTable = []
    for status, user, product, event in results:
        finalEvent = {}
        finalEvent["id"] = event.id
        finalEvent["modelCode"] = product.modelCode
        finalEvent["modelName"] = product.modelName
        finalEvent["username"] = user.username
        finalEvent["statusName"] = status.statusName
        finalEvent["okCounter"] = event.okCounter
        finalEvent["nokCounter"] = event.nokCounter
        finalEvent["startDate"] = datetime.fromtimestamp(event.startDate)
        if event.endDate:
            finalEvent["endDate"] = datetime.fromtimestamp(event.endDate)
            finalEvent["delta"] = round(
                (event.endDate - event.startDate)/3600, 1)
        else:
            finalEvent["endDate"] = "in progress"
            finalEvent["delta"] = "in progress"
        finalEventTable.append(finalEvent)

    # for resfinalEvent in finalEventTable:
    #     print(resfinalEvent)

    if form.validate_on_submit():
        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet("test")
        ws.title = "New Title"
        ws.sheet_properties.tabColor = "1072BA"
        ws['A1'] = "id"
        ws['B1'] = "modelCode"
        ws['C1'] = "modelName"
        ws['D1'] = "username"
        ws['E1'] = "statusName"
        ws['F1'] = "okCounter"
        ws['G1'] = "nokCounter"
        ws['H1'] = "startDate"
        ws['I1'] = "endDate"
        ws['J1'] = "delta[h]"

        # alphabet = list(string.ascii_lowercase)
        # print(alphabet)
        row_counter = 1
        for resfinalEvent in finalEventTable:
            row_counter += 1
            ws['A'+str(row_counter)] = resfinalEvent["id"]
            ws['B'+str(row_counter)] = resfinalEvent["modelCode"]
            ws['C'+str(row_counter)] = resfinalEvent["modelName"]
            ws['D'+str(row_counter)] = resfinalEvent["username"]
            ws['E'+str(row_counter)] = resfinalEvent["statusName"]
            ws['F'+str(row_counter)] = resfinalEvent["okCounter"]
            ws['G'+str(row_counter)] = resfinalEvent["nokCounter"]
            ws['H'+str(row_counter)] = resfinalEvent["startDate"]
            ws['I'+str(row_counter)] = resfinalEvent["endDate"]
            ws['J'+str(row_counter)] = resfinalEvent["delta"]
            # print(row_counter)
        # output_file_name = "output/raport - " + str(date_now) + ".xls"
        output_file_name = "output/raport.xls"
        wb.save(output_file_name)
        flash(
            f'Export complete! File name: {output_file_name}', category='success')
    openEvents = openEventsCounter()

    return render_template('eventTable.html', finalEventTable=finalEventTable, form=form, endDateTimestamp=endDateTimestamp, startDateTimestamp=startDateTimestamp, openEvents=openEvents)
